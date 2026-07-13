"""Разбор файла «Расписание *.xlsx» (по педагогам) → JSON и Excel для справочников АРМ."""
from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

from schedule_markdown import write_markdown_guide

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "Расписание (1)-1-1.xlsx"
DAY_START_COLS = list(range(0, 28, 4))
WEEKDAY_COLS = DAY_START_COLS[:5]  # пн–пт, без сб/вс

PEDAGOG_HEADER = re.compile(r"^\s*Педагог:\s*(.+?)\s*$", re.I | re.U)
LESSON_NUM = re.compile(r"^\d+$")
TIME_RE = re.compile(r"^\d{1,2}:\d{2}$")

COMBINED_SUFFIX = re.compile(
    r"_(\d+)([А-ЯЁ]+)\s*класс",
    re.I | re.U,
)

# «Литература 5 М 2025-26_5ИЛМ классы» → только класс 5М; _5ИЛМ = объединённый урок, не список классов.
EXPLICIT_CLASS_PATTERNS = [
    re.compile(
        r"^(?P<subj>.+?)\s+(?P<grade>\d+)\s+(?P<letter>[А-ЯЁ])\s+(?:\d{4}[-/]|_)",
        re.I | re.U,
    ),
    re.compile(
        r"^(?P<subj>.+?)\s+(?P<grade>\d+)-(?P<letter>[А-ЯЁ])\s+(?:\d{4}[-/]|_)",
        re.I | re.U,
    ),
]

# Короткие подписи без года: «Англ.яз. 5И Комагоркина», «Англ 7 Б Шишакина», «Англ. яз_2 Г_…».
COMPACT_CLASS_PATTERNS = [
    re.compile(
        r"^(?P<subj>.+?)\s+(?P<grade>\d+)(?P<letter>[А-ЯЁ])\s+[А-ЯЁ]",
        re.I | re.U,
    ),
    re.compile(
        r"^(?P<subj>.+?)[_\s]+(?P<grade>\d+)\s*(?P<letter>[А-ЯЁ])_",
        re.I | re.U,
    ),
    re.compile(
        r"^(?P<subj>.+?)\s+(?P<grade>\d+)\s+(?P<letter>[А-ЯЁ])\s+[А-ЯЁ][а-яё]",
        re.I | re.U,
    ),
]

NDO_CLASS_PATTERN = re.compile(
    r"^(?P<subj>.+?)\s+(?P<grade>\d+)\s+НДО\s+",
    re.I | re.U,
)


def lesson_group_signature(top: str, class_key: str) -> tuple:
    return ("single", class_key)

SUBJECT_CANONICAL = {
    "англ": "Английский язык",
    "англ.": "Английский язык",
    "англ.яз": "Английский язык",
    "англ. яз": "Английский язык",
    "рус.яз": "Русский язык",
    "рус. яз": "Русский язык",
    "мат": "Математика",
    "алг": "Алгебра",
    "геом": "Геометрия",
    "физ-ра": "Физическая культура",
    "физкульт": "Физическая культура",
    "физкультура": "Физическая культура",
    "физическаякультура": "Физическая культура",
    "физ": "Физика",
    "ист": "История",
    "общ": "Обществознание",
    "био": "Биология",
    "хим": "Химия",
    "геог": "География",
    "инф": "Информатика",
    "лит": "Литература",
    "муз": "Музыка",
    "изо": "Изобразительное искусство",
    "тех": "Технология",
    "обж": "ОБЖ",
}

CANONICAL_PREFIXES = sorted(SUBJECT_CANONICAL.items(), key=lambda x: len(x[0]), reverse=True)

CLASS_PATTERNS = [
    re.compile(r"^(?P<subj>.+?)\s+(?P<grade>\d+)\s+(?P<letter>[А-ЯЁ])\s", re.I | re.U),
    re.compile(r"^(?P<subj>.+?)\s+(?P<grade>\d+)\s+(?P<letter>[А-ЯЁ])\s+\d{4}", re.I | re.U),
    re.compile(r"^(?P<subj>.+?)\s+(?P<grade>\d+)\s+(?P<letter>[А-ЯЁ])\s+[А-ЯЁ][а-яё]", re.I | re.U),
    re.compile(r"^(?P<subj>.+?)\s+(?P<grade>\d+)\s+(?P<letter>[А-ЯЁ])$", re.I | re.U),
    re.compile(r"^(?P<subj>.+?)_\s*(?P<grade>\d+)\s*(?P<letter>[А-ЯЁ])_", re.I | re.U),
    re.compile(r"^(?P<subj>.+?)\s+(?P<grade>\d+)\s*[кk]\.?\s*(?P<letter>[А-ЯЁ])", re.I | re.U),
]


def cell_str(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    if isinstance(v, pd.Timestamp):
        return v.strftime("%H:%M")
    return str(v).strip()


def canonical_building(line: str) -> str:
    low = line.lower()
    if "основное здание" in low:
        return "Основное здание"
    if any(x in low for x in ("белочка", "здание 8", "пионерская", "начальная школа «белочка»")):
        return "Начальная школа «Белочка»"
    if "литературно" in low or "лмк" in low:
        return "ЛМК"
    if "флигель" in low:
        return "Флигель"
    return "Основное здание"


def parse_room_line(line: str) -> tuple[str, str]:
    line = line.strip()
    if not line:
        return "", ""
    room = line.split("(", 1)[0].strip() if "(" in line else line.split(",", 1)[0].strip()
    return room, canonical_building(line)


def canonical_subject(raw: str) -> str:
    s = raw.strip()
    low = s.lower().replace("ё", "е")
    if any(
        x in low
        for x in (
            "физическая культура",
            "физ-ра",
            "физкульт",
            "физ культура",
            "физ. культура",
        )
    ):
        return "Физическая культура"
    s = re.sub(r"\s+\d+\s+[А-ЯЁ]\s+.*$", "", s, flags=re.U)
    s = re.sub(r"[_\s]\d+[А-ЯЁ]+\s*класс.*$", "", s, flags=re.I | re.U)
    s = re.sub(r"\s+\d{4}[-/].*$", "", s)
    s = re.sub(r"_\d+.*$", "", s)
    s = re.sub(r"\s+[А-ЯЁ]\.\s*[А-ЯЁ]\.?\s*$", "", s)
    s = re.sub(r"\s+[А-ЯЁ]{2,}\s*$", "", s)
    s = s.replace("_", " ").strip(" .")
    key = re.sub(r"[\s.\-]", "", s.lower())
    if key.startswith("физическаякультура") or key.startswith("физкульт"):
        return "Физическая культура"
    for prefix, full in CANONICAL_PREFIXES:
        pk = re.sub(r"[\s.\-]", "", prefix.lower())
        if key.startswith(pk):
            return full
    if key.startswith("англ"):
        return "Английский язык"
    return s[:1].upper() + s[1:] if s else s


def _is_combined_lesson(top: str, letter: str) -> bool:
    m = COMBINED_SUFFIX.search(top)
    if not m:
        return False
    suffix = m.group(2).upper()
    return len(suffix) > 1 or (len(suffix) == 1 and suffix != letter.upper())


def parse_class_subject(top: str) -> list[tuple[str, str, bool]]:
    """Один класс на ячейку. Суффикс _5ИЛМ — признак объединения, не размножение часов."""
    top = top.strip()
    if not top:
        return []

    ndo = NDO_CLASS_PATTERN.match(top)
    if ndo:
        grade = ndo.group("grade")
        return [(f"{grade}НДО", canonical_subject(ndo.group("subj")), True)]

    all_patterns = EXPLICIT_CLASS_PATTERNS + COMPACT_CLASS_PATTERNS + CLASS_PATTERNS
    for pat in all_patterns:
        m = pat.match(top)
        if not m:
            continue
        subj = canonical_subject(m.group("subj"))
        grade = m.group("grade")
        letter = m.group("letter").upper()
        sg = bool(
            re.search(r"группа|п/г|подгруп", top, re.I)
            or _is_combined_lesson(top, letter)
        )
        return [(f"{grade}{letter}", subj, sg)]
    return []


def parse_lesson_pair(subj_cell: str, room_cell: str) -> list[dict]:
    subj_cell = subj_cell.replace("\r\n", "\n").strip()
    room_cell = room_cell.replace("\r\n", "\n").strip()
    if not subj_cell:
        return []
    top = subj_cell.split("\n")[0].strip()
    room_line = room_cell.split("\n")[-1] if room_cell else ""
    if not room_line and "\n" in subj_cell:
        room_line = subj_cell.split("\n", 1)[1].strip()
    room, building = parse_room_line(room_line)
    out = []
    for class_key, subject, has_sg in parse_class_subject(top):
        out.append(
            {
                "class": class_key,
                "subject": subject,
                "room": room,
                "building": building,
                "has_subgroup": has_sg,
            }
        )
    return out


def scan_workbook(path: Path) -> dict:
    df = pd.read_excel(path, sheet_name=0, header=None, engine="openpyxl")
    nrows = df.shape[0]

    teachers: dict[str, Counter] = defaultdict(Counter)
    teacher_lessons: Counter[str] = Counter()
    teacher_class_subject: Counter[tuple[str, str, str]] = Counter()
    teacher_names: list[str] = []
    current_teacher: str | None = None
    curriculum: Counter[tuple[str, str, str]] = Counter()
    raw_events: list[dict] = []
    buildings: set[str] = set()
    rooms: set[tuple[str, str]] = set()
    subjects: set[str] = set()
    classes_meta: dict[str, dict] = {}
    bell_votes: Counter[tuple[int, int, str]] = Counter()

    for r in range(nrows):
        c0 = cell_str(df.iloc[r, 0])
        m = PEDAGOG_HEADER.match(c0)
        if m:
            current_teacher = m.group(1).strip()
            if current_teacher not in teacher_names:
                teacher_names.append(current_teacher)
            continue

        for base in WEEKDAY_COLS:
            lesson_raw = cell_str(df.iloc[r, base])
            if not LESSON_NUM.match(lesson_raw):
                continue
            lesson_num = int(lesson_raw)
            lesson_time = cell_str(df.iloc[r, base + 1])
            if TIME_RE.match(lesson_time):
                shift = 2 if lesson_time >= "12:00" else 1
                bell_votes[(shift, lesson_num, lesson_time)] += 1

            subj = cell_str(df.iloc[r, base + 2])
            room = cell_str(df.iloc[r, base + 3])
            top_line = subj.replace("\r\n", "\n").split("\n")[0].strip()
            parsed = parse_lesson_pair(subj, room)
            if not parsed or not current_teacher:
                continue
            classes = sorted({p["class"] for p in parsed})
            subj_name = parsed[0]["subject"]
            raw_events.append(
                {
                    "teacher": current_teacher,
                    "day": base,
                    "lesson": lesson_num,
                    "subject": subj_name,
                    "classes": classes,
                    "parsed": parsed,
                    "top": top_line,
                    "has_subgroup": any(p["has_subgroup"] for p in parsed),
                    "lesson_time": lesson_time,
                }
            )
            for item in parsed:
                cl = item["class"]
                subjects.add(subj_name)
                teachers[current_teacher][subj_name] += 1
                teacher_class_subject[(current_teacher, cl, subj_name)] += 1
                if item["building"]:
                    buildings.add(item["building"])
                if item["room"] and item["building"]:
                    rooms.add((item["building"], item["room"]))
                grade = int(re.match(r"(\d+)", cl).group(1))
                meta = classes_meta.setdefault(
                    cl,
                    {"grade": grade, "letter": cl[len(str(grade)) :], "shift_guess": set()},
                )
                if TIME_RE.match(lesson_time):
                    meta["shift_guess"].add(2 if lesson_time >= "12:00" else 1)

    lesson_events: set[tuple] = set()
    for event in raw_events:
        subj_name = event["subject"]
        sig = lesson_group_signature(event["top"], event["classes"][0])
        event_key = (event["day"], event["lesson"], subj_name, sig)
        if event_key in lesson_events:
            continue
        lesson_events.add(event_key)
        sg = "да" if event["has_subgroup"] else "нет"
        teacher_lessons[event["teacher"]] += 1
        for item in event["parsed"]:
            cl = item["class"]
            curriculum[(cl, subj_name, sg)] += 1

    classes_rows = []
    for cl in sorted(classes_meta, key=lambda x: (classes_meta[x]["grade"], classes_meta[x]["letter"])):
        meta = classes_meta[cl]
        shifts = meta["shift_guess"]
        shift = shifts.pop() if len(shifts) == 1 else 1
        classes_rows.append(
            {
                "grade": meta["grade"],
                "letter": meta["letter"],
                "display": cl,
                "shift": shift,
            }
        )

    bells_canonical = []
    for shift in (1, 2):
        lessons = sorted({k[1] for k in bell_votes if k[0] == shift})
        for lesson in lessons:
            times = [
                (t, c)
                for (sh, les, t), c in bell_votes.items()
                if sh == shift and les == lesson
            ]
            if not times:
                continue
            best_time = max(times, key=lambda x: x[1])[0]
            bells_canonical.append(
                {
                    "template": "1 смена" if shift == 1 else "2 смена",
                    "shift": shift,
                    "lesson": lesson,
                    "start": best_time,
                }
            )

    def teacher_hint(class_key: str, subject: str) -> str:
        matches = [
            (t, c)
            for (t, cl, subj), c in teacher_class_subject.items()
            if cl == class_key and subj == subject
        ]
        if not matches:
            return "—"
        return max(matches, key=lambda x: x[1])[0]

    curriculum_rows = [
        {
            "class": cl,
            "subject": subj,
            "hours_per_week": h,
            "subgroups": sg,
            "teacher_hint": teacher_hint(cl, subj),
        }
        for (cl, subj, sg), h in curriculum.items()
    ]

    return {
        "source": path.name,
        "school_hint": "МБОУ СОШ №14 (по адресам в файле)",
        "teachers": teacher_names,
        "buildings": sorted(buildings),
        "rooms": [{"building": b, "number": n} for b, n in sorted(rooms)],
        "subjects": sorted(subjects),
        "classes": classes_rows,
        "curriculum": sorted(
            curriculum_rows,
            key=lambda x: (
                classes_meta.get(x["class"], {"grade": 99})["grade"],
                x["class"],
                x["subject"],
            ),
        ),
        "bells": bells_canonical,
        "teacher_subjects": [
            {
                "name": name,
                "subjects": dict(teachers[name]),
                "lessons_per_week": teacher_lessons[name],
            }
            for name in teacher_names
        ],
    }


def write_excel_template(data: dict, out_path: Path) -> None:
    wb = Workbook()
    bold = Font(bold=True)

    def sheet(name: str, headers: list[str], rows: list[list]):
        ws = wb.create_sheet(name)
        for i, h in enumerate(headers, 1):
            c = ws.cell(1, i, h)
            c.font = bold
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                ws.cell(ri, ci, val)

    wb.remove(wb.active)
    sheet("Здания", ["Название", "Цвет_HEX"], [[b, ""] for b in data["buildings"]])
    sheet(
        "Кабинеты",
        ["Номер", "Здание", "Вместимость", "Специфика", "Закреплённый_учитель"],
        [[r["number"], r["building"], 30, "", ""] for r in data["rooms"]],
    )
    sheet(
        "Предметы",
        ["Название", "Балл_Сивкова"],
        [[s, ""] for s in data["subjects"]],
    )
    sheet(
        "Классы",
        ["Параллель", "Буква", "Смена", "Учеников", "Коррекционный", "Здание"],
        [
            [c["grade"], c["letter"], c["shift"], 25, "", ""]
            for c in data["classes"]
        ],
    )
    sheet(
        "Учителя",
        [
            "ФИО",
            "Тип",
            "Должность",
            "Макс_нагрузка",
            "Основной_профиль",
            "Смежный_профиль",
            "Классное_руководство",
            "Телефон",
            "Контакт_URL",
            "Контакт_заметка",
        ],
        [
            [
                t,
                "Предметник",
                "",
                18,
                max(ts["subjects"], key=ts["subjects"].get) if ts["subjects"] else "",
                "",
                "",
                "",
                "",
                "",
            ]
            for t, ts in zip(data["teachers"], data["teacher_subjects"])
        ],
    )
    sheet(
        "Нагрузка",
        ["Класс", "Предмет", "Часов_в_неделю", "Подгруппы", "Неделя"],
        [
            [row["class"], row["subject"], row["hours_per_week"], row["subgroups"], "каждую"]
            for row in data["curriculum"]
        ],
    )
    bell_rows = []
    for b in data["bells"]:
        bell_rows.append(
            [b["template"], 1, 11, "урок", b["lesson"], b["start"], "", b["shift"]]
        )
    sheet(
        "Звонки",
        ["Шаблон", "Параллель_с", "Параллель_по", "Тип", "Номер_урока", "Начало", "Конец", "Смена"],
        bell_rows,
    )
    wb.save(out_path)


def main() -> None:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    data = scan_workbook(path)
    json_path = path.with_name(path.stem + ".parsed.json")
    json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    xlsx_path = path.with_name("Справочники-из-расписания.xlsx")
    md_path = ROOT / "DOC" / "Справочники-из-расписания.md"
    write_excel_template(data, xlsx_path)
    write_markdown_guide(data, md_path)
    print(
        f"teachers={len(data['teachers'])}, classes={len(data['classes'])}, "
        f"curriculum={len(data['curriculum'])}, rooms={len(data['rooms'])}"
    )
    print(md_path)


if __name__ == "__main__":
    main()
